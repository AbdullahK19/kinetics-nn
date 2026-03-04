"""
noise_robustness.py — Ensemble noise-robustness analysis.

Adds independent zero-mean Gaussian noise at 101 levels (0% to 50% in 0.5% steps)
to the normalised test set and evaluates ensemble accuracy at each level.

Noise definition:
    noisy_X = X_norm + N(0, σ²),  where σ = noise_pct / 100
    (noise std expressed as fraction of normalised feature std ≈ 1)

Outputs (saved to results/):
    noise_robustness.xlsx  — Excel workbook with three sheets:
        "Overall"          — noise_pct, accuracy_pct, n_correct, n_total
        "Per_Class_Recall" — noise_pct × 28 class recall columns
        "Breakpoints"      — per-class noise level at which recall first drops below 80%

Usage:
    PYTHONPATH=src python3 noise_robustness.py
"""

import json
import sys
import numpy as np
import torch
from pathlib import Path
from tqdm import tqdm

# ── Config ────────────────────────────────────────────────────────────────────
ENSEMBLE_DIRS = ["models/ensemble_1", "models/ensemble_2", "models/ensemble_3"]
DATA_DIR      = Path("data")
RESULTS_DIR   = Path("results")
DEVICE        = torch.device("mps" if torch.backends.mps.is_available() else "cpu")
BATCH_SIZE    = 2000        # samples per forward-pass batch
SEED          = 42          # for reproducible noise generation
RECALL_THRESHOLD = 80.0     # % — used for breakpoint calculation

NOISE_LEVELS  = np.round(np.arange(0.0, 50.5, 0.5), 1)   # 0.0, 0.5, ..., 50.0
# Precompute index lookup for specific noise levels used in summaries
_IDX = {float(v): i for i, v in enumerate(NOISE_LEVELS)}

print(f"Device: {DEVICE}")
print(f"Noise levels: {len(NOISE_LEVELS)} steps ({NOISE_LEVELS[0]}% → {NOISE_LEVELS[-1]}%)")

# ── Load ensemble models ──────────────────────────────────────────────────────
sys.path.insert(0, "src")
from ensemble_predict import load_model_from_dir, load_class_names

RESULTS_DIR.mkdir(exist_ok=True)

print(f"\nLoading {len(ENSEMBLE_DIRS)} ensemble models...")
models, cfgs = [], []
for d in ENSEMBLE_DIRS:
    m, cfg = load_model_from_dir(Path(d), DEVICE)
    models.append(m)
    cfgs.append(cfg)
    print(f"  ✓ {d}  (epoch {cfg.get('best_epoch','?')}, val_acc={cfg.get('best_val_acc',0)*100:.2f}%)")

class_names_dict = load_class_names(DATA_DIR)
class_names = [class_names_dict.get(i, f"Class {i}") for i in range(28)]
n_classes = len(class_names)

# ── Load test set (already normalised) ───────────────────────────────────────
print("\nLoading test set...")
test   = np.load(DATA_DIR / "test.npz")
X_norm = test["X"].astype(np.float32)   # (N, 128, 16) — already z-score normalised
y_test = test["y"].astype(np.int64)
N      = len(y_test)
print(f"  {N} samples, {n_classes} classes  |  feature std ≈ {X_norm.std():.3f}")


# ── Helper: batched ensemble inference ───────────────────────────────────────
@torch.no_grad()
def predict_ensemble(X_np: np.ndarray) -> np.ndarray:
    """Return mean softmax probabilities from ensemble. Shape: (N, n_classes)."""
    all_probs = []
    for model in models:
        probs_batches = []
        for start in range(0, len(X_np), BATCH_SIZE):
            batch = torch.tensor(X_np[start:start + BATCH_SIZE],
                                 dtype=torch.float32).to(DEVICE)
            logits = model(batch)
            probs_batches.append(torch.softmax(logits, dim=-1).cpu().numpy())
        all_probs.append(np.concatenate(probs_batches, axis=0))  # (N, n_classes)
    return np.mean(all_probs, axis=0)   # averaged over models


# ── Main loop ─────────────────────────────────────────────────────────────────
rng = np.random.default_rng(SEED)

rows_overall    = []   # [noise_pct, acc, n_correct, n_total]
rows_per_class  = []   # [noise_pct, recall_c0, ..., recall_c27]

print(f"\nRunning noise analysis ({len(NOISE_LEVELS)} levels)...\n")

for noise_pct in tqdm(NOISE_LEVELS, desc="Noise levels", ncols=80):
    sigma = noise_pct / 100.0

    if sigma == 0.0:
        X_noisy = X_norm
    else:
        noise = rng.standard_normal(X_norm.shape).astype(np.float32) * sigma
        X_noisy = X_norm + noise

    mean_probs = predict_ensemble(X_noisy)
    preds      = mean_probs.argmax(axis=1)

    # Overall accuracy
    n_correct = int((preds == y_test).sum())
    acc       = n_correct / N * 100.0
    rows_overall.append([noise_pct, round(acc, 4), n_correct, N])

    # Per-class recall
    recalls = []
    for c in range(n_classes):
        mask = y_test == c
        recall = (preds[mask] == c).mean() * 100.0 if mask.sum() > 0 else float('nan')
        recalls.append(round(recall, 4))
    rows_per_class.append([noise_pct] + recalls)


# ── Compute breakpoints ───────────────────────────────────────────────────────
# For each class: first noise level where recall drops below RECALL_THRESHOLD %
breakpoints = []
per_class_arr = np.array([row[1:] for row in rows_per_class])  # (101, 28)

for c in range(n_classes):
    recall_curve = per_class_arr[:, c]
    below = np.where(recall_curve < RECALL_THRESHOLD)[0]
    if len(below) == 0:
        bp = f"> {NOISE_LEVELS[-1]}%"
    else:
        bp = f"{NOISE_LEVELS[below[0]]}%"
    breakpoints.append({
        "class_id":   c,
        "class_name": class_names[c],
        "breakpoint": bp,
        "recall_at_0pct":  round(float(recall_curve[0]),  2),
        "recall_at_10pct": round(float(recall_curve[_IDX[10.0]]), 2),
        "recall_at_25pct": round(float(recall_curve[_IDX[25.0]]), 2),
        "recall_at_50pct": round(float(recall_curve[-1]), 2),
    })

# Sort by breakpoint (most fragile first)
def bp_sort_key(b):
    bp = b["breakpoint"]
    if bp.startswith(">"):
        return 999.0
    return float(bp.rstrip("%"))

breakpoints_sorted = sorted(breakpoints, key=bp_sort_key)


# ── Write Excel workbook ──────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

output_path = RESULTS_DIR / "noise_robustness.xlsx"
wb = openpyxl.Workbook()

# ── Sheet 1: Overall ──────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Overall"

headers1 = ["Noise (%)", "Ensemble Accuracy (%)", "Correct", "Total"]
ws1.append(headers1)
for cell in ws1[1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.font = Font(bold=True, color="FFFFFF")

for row in rows_overall:
    ws1.append(row)

# Freeze header row, auto-width
ws1.freeze_panes = "A2"
for col in range(1, 5):
    ws1.column_dimensions[get_column_letter(col)].width = 22

# ── Sheet 2: Per-class recall ─────────────────────────────────────────────────
ws2 = wb.create_sheet("Per_Class_Recall")

headers2 = ["Noise (%)"] + [f"[{i}] {class_names[i]}" for i in range(n_classes)]
ws2.append(headers2)
for cell in ws2[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(wrap_text=True)

ws2.row_dimensions[1].height = 50
ws2.freeze_panes = "B2"

for row in rows_per_class:
    ws2.append(row)

# Colour-code recall cells: red=low, green=high
from openpyxl.formatting.rule import ColorScaleRule
data_range = f"B2:{get_column_letter(n_classes + 1)}{len(rows_per_class) + 1}"
ws2.conditional_formatting.add(
    data_range,
    ColorScaleRule(
        start_type="num", start_value=0,   start_color="FF0000",
        mid_type="num",   mid_value=80,    mid_color="FFFF00",
        end_type="num",   end_value=100,   end_color="00B050",
    )
)

for col in range(1, n_classes + 2):
    ws2.column_dimensions[get_column_letter(col)].width = 16
ws2.column_dimensions["A"].width = 12

# ── Sheet 3: Breakpoints ──────────────────────────────────────────────────────
ws3 = wb.create_sheet("Breakpoints")

ws3.append([
    f"Breakpoint = noise level at which recall first drops below {RECALL_THRESHOLD}%",
    "", "", "", "", ""
])
ws3.merge_cells("A1:F1")
ws3["A1"].font = Font(bold=True, italic=True)

headers3 = ["Class ID", "Mechanism", f"Breakpoint (<{RECALL_THRESHOLD}%)",
            "Recall @ 0%", "Recall @ 10%", "Recall @ 25%", "Recall @ 50%"]
ws3.append(headers3)
for cell in ws3[2]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")

for bp in breakpoints_sorted:
    ws3.append([
        bp["class_id"],
        bp["class_name"],
        bp["breakpoint"],
        bp["recall_at_0pct"],
        bp["recall_at_10pct"],
        bp["recall_at_25pct"],
        bp["recall_at_50pct"],
    ])

ws3.column_dimensions["A"].width = 10
ws3.column_dimensions["B"].width = 35
ws3.column_dimensions["C"].width = 22
for col in ["D", "E", "F", "G"]:
    ws3.column_dimensions[col].width = 16

# ── Summary stats on Overall sheet ───────────────────────────────────────────
accs = [row[1] for row in rows_overall]
ws1.append([])
ws1.append(["--- Summary ---"])
ws1.append(["Accuracy at 0% noise",  f"{accs[0]:.2f}%"])
ws1.append(["Accuracy at 10% noise", f"{accs[_IDX[10.0]]:.2f}%"])
ws1.append(["Accuracy at 25% noise", f"{accs[_IDX[25.0]]:.2f}%"])
ws1.append(["Accuracy at 50% noise", f"{accs[-1]:.2f}%"])
below50 = [NOISE_LEVELS[i] for i, a in enumerate(accs) if a < 50.0]
ws1.append(["Model breaks (acc < 50%)", f"{below50[0]}%" if below50 else f"> {NOISE_LEVELS[-1]}%"])

wb.save(output_path)
print(f"\n✓ Saved: {output_path}")

# ── Print summary to console ──────────────────────────────────────────────────
print(f"\n{'='*55}")
print(f"  NOISE ROBUSTNESS SUMMARY")
print(f"{'='*55}")
print(f"  Accuracy @ 0%  noise: {accs[0]:.2f}%")
print(f"  Accuracy @ 5%  noise: {accs[_IDX[5.0]]:.2f}%")
print(f"  Accuracy @ 10% noise: {accs[_IDX[10.0]]:.2f}%")
print(f"  Accuracy @ 25% noise: {accs[_IDX[25.0]]:.2f}%")
print(f"  Accuracy @ 50% noise: {accs[-1]:.2f}%")
print(f"\n  Most robust classes (breakpoint > 25%):")
for bp in reversed(breakpoints_sorted):
    val = bp_sort_key(bp)
    if val > 25.0:
        print(f"    [{bp['class_id']:2d}] {bp['class_name']:<35} → {bp['breakpoint']}")
print(f"\n  Most fragile classes (breakpoint ≤ 10%):")
for bp in breakpoints_sorted:
    val = bp_sort_key(bp)
    if val <= 10.0:
        print(f"    [{bp['class_id']:2d}] {bp['class_name']:<35} → {bp['breakpoint']}")
print(f"{'='*55}")
